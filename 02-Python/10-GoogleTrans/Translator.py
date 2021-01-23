# /usr/bin/env python
# coding=utf-8

import os
import sys
import xlrd
from openpyxl import load_workbook
# from googletrans import Translator
from google_trans_new import google_translator

import time

# translator = Translator(['translate.google.cn'])
translator = google_translator()
detector = google_translator()

# Function: Find the specific column
def findtargetcolumn(columnheaders, SubjectCol, LanguageCol):
    global subjects_Text, Subject_col, languages_code, language_col
    column = 0
    flag = 0
    for head in columnheaders:
        if head == SubjectCol:
            subjects_Text = sheet.col_values(column)
            Subject_col = column + 1
            flag = 1

        elif head == LanguageCol:
            languages_code = sheet.col_values(column)
            language_col = column + 1

        column = column + 1

    return subjects_Text, Subject_col, languages_code, language_col, flag

# Function: Do translation and write into Excel file.
# Last updated on 23th,Jan, 2021
# Creat Repository for Git and Edit with VSCode

# To be Updated:
'''
1. Modulize the Function
2. Allow long text > 5000
3. TBD...
'''

def google_translation(subjects_Text, Subject_col, languages_code, targetsheet, workbook):

    del subjects_Text[0]    # Remove the first row of Title: Subject, etc
    del languages_code[0]
    subject_row = 1     # Start from Row 2 to skip the first row of Headline.

    targetsheet.cell(row=1, column=Subject_col).value = "Subject Before Translation"
    targetsheet.cell(row=1, column=Subject_col+1).value = "Original Language Code"

    targetsheet.cell(row=1, column=Subject_col+2).value = "Subject Translated"
    targetsheet.cell(row=1, column=Subject_col+3).value = "Language Detected"
    targetsheet.cell(row=1, column=Subject_col+4).value = "Translation Status"

    ShortInterval_Avoid_IPBlock = 30  # MaYS added for avoid IP Block
    LongInterval_Avoid_IPBlock = 100  # MaYS added for avoid IP Block
    
    for i in range(0, len(subjects_Text)):
    #for subjects_Text in subjects_Text:

        # Request interval time, default 5 sec, may need to be changed to longer to avoid been IP blocked.
        time.sleep(5)
        subject_row = subject_row + 1

        # MaYS added for avoid IP Block
        ShortInterval_Avoid_IPBlock = ShortInterval_Avoid_IPBlock - 1
        LongInterval_Avoid_IPBlock = LongInterval_Avoid_IPBlock - 1

        if ShortInterval_Avoid_IPBlock == 0:
            print(
                'Waiting for 40s after 30 requests to avoid being blocked')
            time.sleep(80)
            ShortInterval_Avoid_IPBlock = 30

        if LongInterval_Avoid_IPBlock == 0:
            print(
                'Waiting for 90s after 100 requests to avoid being blocked')
            time.sleep(180)
            LongInterval_Avoid_IPBlock = 100
            # MaYS added for avoid IP Block

        if (subjects_Text[i] != '') & (subjects_Text[i] != '#N/A'):
            try:
                '''
                if ((translator.detect(case_act).lang != 'en') | (
                  (translator.detect(case_act).lang == 'en') & (
                      translator.detect(case_act).confidence < 0.98))):
                # Obsoleted for the old lib of GoogleTrans does not work for tk Value
                '''
                language_code = ""

                if languages_code[i] == "":
                    
                    subject_text_translated = translator.translate(subjects_Text[i], lang_tgt='en')
                    languageDetected = detector.detect(subjects_Text[i])
                    language_code = languageDetected[1]

                subject_text_translated = translator.translate(
                    subjects_Text[i], lang_tgt='en', lang_src=languages_code[i])

                # write the raw text before Translation in the Column 1

                targetsheet.cell(
                    row=subject_row, column=Subject_col).value = subjects_Text[i]
                targetsheet.cell(
                    row=subject_row, column=Subject_col+1).value = languages_code[i]

                targetsheet.cell(
                    row=subject_row, column=Subject_col+2).value = subject_text_translated
                targetsheet.cell(
                    row=subject_row, column=Subject_col+3).value = language_code
                targetsheet.cell(
                    row=subject_row, column=Subject_col+4).value = "Translation Success"

                print('Line' + str(subject_row) + '\tTranslation Done!')

                workbook.save("Raw Data_Translated.xlsx")

                # MaYS added for avoid IP Block

            except(Exception):

                targetsheet.cell(
                    row=subject_row, column=Subject_col).value = subjects_Text[i]
                targetsheet.cell(
                    row=subject_row, column=Subject_col + 1).value = languages_code[i]

                targetsheet.cell(
                    row=subject_row, column=Subject_col+4).value = "Translation Error"

                print('Line' + str(subject_row) + '\tTranslated Error!')

                workbook.save("Raw Data_Translated.xlsx")

    try:
        print('Translation completed!')
        workbook.save("Raw Data_Translated.xlsx")
    except(Exception):
        print('Save Error!')


# Load initial data from Excel file, do translation and save the translation file.
currentpath = sys.path[0]
if not os.path.exists(currentpath + "\\Raw Data.xlsx"):
    print("No available data.")

else:
    workbook_read = xlrd.open_workbook(currentpath + "\\Raw Data.xlsx")
    print("Read Data Successfully!")
    sheet = workbook_read.sheet_by_name('Raw Data')      # Find "Raw Data" sheet
    headers = sheet.row_values(0)                        # Get headers

    result = findtargetcolumn(headers, 'Subject', 'Language Code')

    if result[4] == 1:
        print("Find \"Subject\" Column!")
        workbook = load_workbook(currentpath + "\\Raw Data.xlsx")
        targetsheet = workbook.get_sheet_by_name('Raw Data')

        targetsheet.title = "Text Translated"
        workbook.save("Raw Data_Translated.xlsx")

        google_translation(result[0], result[1], result[2], targetsheet, workbook)

    else:
        print("No \"Subject\" Column!")

    if os.path.exists(currentpath + "\\Raw Data_Translated.xlsx"):
        try:
            result = findtargetcolumn(headers, 'T2 Activity')
            if result[4] == 1:
                print("Find \"T2 Activity\" Column!")
                workbook = load_workbook(
                    currentpath + "\\Raw Data_Translated.xlsx")
                targetsheet = workbook.get_sheet_by_name('Raw Data')
                targetsheet.title = "Text Translated"
                workbook.save("Raw Data_Translated.xlsx")

                google_translation(result[0], result[1], result[2], targetsheet, workbook)
        except(Exception):
            print("No \"T2 Activity\" Column!")
    else:
        try:
            result = findtargetcolumn(headers, 'T2 Activity')
            if result[4] == 1:
                print("Find \"T2 Activity\" Column!")
                workbook = load_workbook(currentpath + "\\Raw Data.xlsx")
                targetsheet = workbook.get_sheet_by_name('Raw Data')
                google_translation(
                    result[0], result[1], targetsheet, workbook)
        except(Exception):
            print("No \"T2 Activity\" Column!")

    print("Saved file successfully!")
