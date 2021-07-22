# /usr/bin/env python
# coding=utf-8

import os
import re
import string
import sys
import time
import math


import xlrd
# from googletrans import Translator
from google_trans_new import google_translator
from openpyxl import load_workbook

# translator = Translator(['translate.google.cn'])
translator = google_translator()
detector = google_translator()
MAX_CHAR_LIMIT_GOOGLE_TRANS = 4800
# Actually 5000 for Char Limit per Request of Google Translation

# Function: Find the specific column

def findtargetcolumn(_column_headers, _subject_col, _language_col):
    global subjects_Text, Subject_col, languages_code, language_col
    column = 0
    flag = False
    for head in _column_headers:
        if head == _subject_col:
            subjects_Text = sheet.col_values(column)
            Subject_col = column + 1
            flag = True

        elif head == _language_col:
            languages_code = sheet.col_values(column)
            language_col = column + 1

        column = column + 1

    return subjects_Text, Subject_col, languages_code, language_col, flag


# Function: Do translation and write into Excel file.
def AutoTransExcel_Google(_subjects_text, _subject_col, _languages_code, _target_sheet, _workbook):
    del _subjects_text[0]  # Remove the first row of Title: Subject, etc
    del _languages_code[0]
    subject_row = 1  # Start from Row 2 to skip the first row of Headline.

    _target_sheet.cell(row=1, column=_subject_col).value = "Subject Before Translation"
    _target_sheet.cell(row=1, column=_subject_col + 1).value = "Original Language Code"

    _target_sheet.cell(row=1, column=_subject_col + 2).value = "Subject Translated"
    _target_sheet.cell(row=1, column=_subject_col + 3).value = "Language Detected"
    _target_sheet.cell(row=1, column=_subject_col + 4).value = "Translation Status"

    ShortInterval_Avoid_IPBlock = 30  # MaYS added for avoid IP Block
    LongInterval_Avoid_IPBlock = 100  # MaYS added for avoid IP Block

    for i in range(0, len(_subjects_text)):
        # for subjects_Text in subjects_Text:

        # Request interval time, default 5 sec, may need to be changed to longer to avoid been IP blocked.
        time.sleep(5)
        subject_row = subject_row + 1

        # MaYS added for avoid IP Block
        ShortInterval_Avoid_IPBlock = ShortInterval_Avoid_IPBlock - 1
        LongInterval_Avoid_IPBlock = LongInterval_Avoid_IPBlock - 1

        if ShortInterval_Avoid_IPBlock <= 0:
            print(
                'Waiting for 40s after 30 requests to avoid being blocked')
            time.sleep(40)
            ShortInterval_Avoid_IPBlock = 30

        if LongInterval_Avoid_IPBlock <= 0:
            print(
                'Waiting for 90s after 100 requests to avoid being blocked')
            time.sleep(90)
            LongInterval_Avoid_IPBlock = 100
            # MaYS added for avoid IP Block

        if (_subjects_text[i] != '') & (_subjects_text[i] != '#N/A'):
            try:

                language_code = languages_code[i]
                if language_code == "":
                    languageDetected = detector.detect(_subjects_text[i])
                    language_code = languageDetected[1]
                # if Language Code not defined, Use Detector to detect the Language Code

                subject_text_translated = ''

                if (len(_subjects_text[i]) > MAX_CHAR_LIMIT_GOOGLE_TRANS):

                    nNrSegmentLongText = math.ceil(len(_subjects_text[i]) / MAX_CHAR_LIMIT_GOOGLE_TRANS)

                    SegmentedTextTranslated = []

                    LongTextSegmented = re.findall('.{1,4800}',
                                                   _subjects_text[i], re.DOTALL|re.MULTILINE)

                    for j in range(0, nNrSegmentLongText):
                        LongTextTranslated = translator.translate(
                            LongTextSegmented[j], lang_tgt='en', lang_src=language_code)

                        ShortInterval_Avoid_IPBlock = ShortInterval_Avoid_IPBlock - 1
                        LongInterval_Avoid_IPBlock = LongInterval_Avoid_IPBlock - 1
                        
                        time.sleep(8)
                        SegmentedTextTranslated.append(LongTextTranslated)
                        pass  #Reserved for Breakpoint for Debugging

                    subject_text_translated = "".join(SegmentedTextTranslated)
                    pass #Reserved for Breakpoint for Debugging

                    LongTextSegmented.clear()
                    SegmentedTextTranslated.clear()
                else:
                    pass  #Reserved for Breakpoint for Debugging
                    subject_text_translated = translator.translate(
                        _subjects_text[i], lang_tgt='en', lang_src=language_code)

                pass
                # write the raw text before Translation in the Column 1
                _target_sheet.cell(
                    row=subject_row, column=_subject_col).value = _subjects_text[i]
                _target_sheet.cell(
                    row=subject_row, column=_subject_col + 1).value = _languages_code[i]

                _target_sheet.cell(
                    row=subject_row, column=_subject_col + 2).value = subject_text_translated
                _target_sheet.cell(
                    row=subject_row, column=_subject_col + 3).value = language_code
                _target_sheet.cell(
                    row=subject_row, column=_subject_col + 4).value = "Translation Success"

                print('Line' + str(subject_row) + '\tTranslation Done!')

                _workbook.save("ForGoogleTrans_Translated.xlsx")


            except(Exception):

                _target_sheet.cell(
                    row=subject_row, column=_subject_col).value = _subjects_text[i]
                _target_sheet.cell(
                    row=subject_row, column=_subject_col + 1).value = _languages_code[i]

                _target_sheet.cell(
                    row=subject_row, column=_subject_col + 4).value = "Translation Error"

                print('Line' + str(subject_row) + '\tTranslated Error!')

                _workbook.save("ForGoogleTrans_Translated.xlsx")

    try:
        print('Translation completed!')
        _workbook.save("ForGoogleTrans_Translated.xlsx")
    except(Exception):
        print('Save Error!')


# Load initial data from Excel file, do translation and save the translation file.
currentpath = sys.path[0]
if not os.path.exists(currentpath + "\\ForGoogleTrans_SourceText.xlsx"):
    print("No available data.")

else:
    workbook_read = xlrd.open_workbook(currentpath + "\\ForGoogleTrans_SourceText.xlsx")
    print("Read Data Successfully!")
    sheet = workbook_read.sheet_by_name('SourceTextForTrans')  # Find "SourceTextForTrans" sheet
    headers = sheet.row_values(0)  # Get headers

    result = findtargetcolumn(headers, 'Subject', 'Language Code')

    if result[4] == True:
        print("Find \"Subject\" Column!")
        workbook = load_workbook(currentpath + "\\ForGoogleTrans_SourceText.xlsx")
        targetsheet = workbook.get_sheet_by_name('SourceTextForTrans')
        '''
        DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
          targetsheet = workbook.get_sheet_by_name('SourceTextForTrans')
        '''

        targetsheet.title = "Text Translated"
        workbook.save("ForGoogleTrans_Translated.xlsx")

        AutoTransExcel_Google(result[0], result[1], result[2], targetsheet, workbook)

    else:
        print("No \"Subject\" Column!")

    if os.path.exists(currentpath + "\\ForGoogleTrans_Translated.xlsx"):
        try:
            result = findtargetcolumn(headers, 'T2 Activity')
            if result[4] == True:
                print("Find \"T2 Activity\" Column!")
                workbook = load_workbook(
                    currentpath + "\\ForGoogleTrans_Translated.xlsx")
                targetsheet = workbook.get_sheet_by_name('SourceTextForTrans')
                targetsheet.title = "Text Translated"
                workbook.save("ForGoogleTrans_Translated.xlsx")

                AutoTransExcel_Google(result[0], result[1], result[2], targetsheet, workbook)
        except(Exception):
            print("No \"T2 Activity\" Column!")
    else:
        try:
            result = findtargetcolumn(headers, 'T2 Activity')
            if result[4] == True:
                print("Find \"T2 Activity\" Column!")
                workbook = load_workbook(currentpath + "\\ForGoogleTrans_SourceText.xlsx")
                targetsheet = workbook.get_sheet_by_name('SourceTextForTrans')
                AutoTransExcel_Google(
                    result[0], result[1], targetsheet, workbook)
        except(Exception):
            print("No \"T2 Activity\" Column!")

    print("Saved file successfully!")
